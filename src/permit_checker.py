import openpyxl
from openpyxl.utils import column_index_from_string

class PermitChecker:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.workbook = openpyxl.load_workbook(excel_path)
        self.sheet = self.workbook.active

    def get_permit_list(self):
        permit_list = []
        for row in self.sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            permit_number = row[0]
            if permit_number:
                permit_list.append(str(permit_number))
        return permit_list

    def update_permit_status(self, permit_number, status_info):
        for row in self.sheet.iter_rows(min_row=2, max_col=1):
            cell = row[0]
            if cell.value == permit_number:
                for key, value in status_info.items():
                    col_idx = column_index_from_string(key)
                    self.sheet.cell(row=cell.row, column=col_idx, value=value)
                break
        self.workbook.save(self.excel_path)

    def save(self):
        self.workbook.save(self.excel_path)
